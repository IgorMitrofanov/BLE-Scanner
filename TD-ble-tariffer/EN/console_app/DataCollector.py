import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import datetime
import numpy as np


class DataCollector:
    """
    Description of the Data Collector class.

    Attributes:
    - start_serial (int): the initial serial number.
    - end_serial (int): the final serial number.
    - device_names (list): a list of device names.
    - start_len (int): the initial length of the device list.
    - device_type (str): device type.

    Methods:
    - __init__(self, start_serial, end_serial, device_type): class constructor.
    - update_char(self, device_name, char_name, value): method for updating the characteristic value.
    - - get_data frame(self): a method for getting data in the form of a pandas DataFrame.
    - to_excel(self, xls_path): a method for saving data to an Excel file.
    """
    def __init__(self, start_serial, end_serial, device_type):
        """
        Constructor of the Data Collector class.

        Parameters:
        - start_serial (int): the initial serial number.
        - end_serial (int): the final serial number.
        - device_type (str): device type.
        """
        self.start_serial = start_serial
        self.end_serial = end_serial
        self.device_names = [device_type + '_' + str(i) for i in range(start_serial, end_serial+1)]
        self.start_len = len(self.device_names)
        self.device_type = device_type

        if device_type == 'TD':

            self.df = pd.DataFrame({'Name': self.device_names, 
                                    'MAC': [None] * len(self.device_names),
                                    'RSSI': [None] * len(self.device_names),
                                    'version': [None] * len(self.device_names),
                                    'Battery voltage': [None] * len(self.device_names),
                                    'Temperature': [None] * len(self.device_names),
                                    'Fuel level' : [None] * len(self.device_names),
                                    'Period': [None] * len(self.device_names),
                                    'H': [None] * len(self.device_names),
                                    'L': [None] * len(self.device_names),
                                    'Fuel level after calibrate': [None] * len(self.device_names),
                                    'Why rejected': [None] * len(self.device_names)})


    def update_char(self, device_name, char_name, value):
        """
        Method for updating the characteristic value.

        Parameters:
        - device_name (str): device name.
        - char_name (str): the name of the characteristic.
        - value: the new value of the characteristic.
        """
        index = self.df.index[self.df['Name'] == device_name][0]
        self.df.loc[index, char_name] = value
            

    def get_dataframe(self):
        """
        Method for getting data in the form of pandas DataFrame.

        Returns:
        - pandas.DataFrame: data in the form of a DataFrame.
        """
        return self.df
    

    def to_excel(self, xls_path, atrribute_error_flag):
        """
        Method for saving data to an Excel file.

        Parameters:
        - xls_path (str): path to the Excel file.
        - - attribute_error_flag (bool): a crutch for the bleak library, allows you to make a note that not all devices have been calibrated
        """
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H_%M')
        self.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active

        # Defining the fill color
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        white_font = Font(color='FFFFFF') # White font color

        avg_temp = np.mean([row[6].value for row in ws.iter_rows(min_row=2)])

        # We go through all the lines, starting from the second
        for row in ws.iter_rows(min_row=2):
            temp = row[6].value 
            hk = row[9].value 
            lk = row[10].value 
            fl = row[11].value
            if hk and lk and fl is not None:
                if fl > 15:
                    row[12].value = "The fuel level is more than 15 units"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif lk < 20000:
                    row[12].value = "L < 20000"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif hk > 43000:
                    row[12].value = "L < 20000"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif abs(temp - avg_temp) > 5:
                    row[12].value = "The temperature differs from the average by more than 5 units"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
        if atrribute_error_flag == True:
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1).value = "SOME DEVICES COULD NOT COMPLETE CALIBRATION, THE PROGRAM NEEDS TO BE RESTARTED"
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[last_row]:
                cell.fill = red_fill
                cell.font = font
        wb.save(xls_path)

