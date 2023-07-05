import os

try:
    import pandas as pd
except:
    os.system('pip install pandas==2.0.2')
    import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except:
    os.system('pip install openpyxl==3.1.2')
    import openpyxl
    from openpyxl.styles import PatternFill, Font


import datetime
import numpy as np

class DataCollector:
    def __init__(self, start_serial, end_serial, device_type):
        self.start_serial = start_serial
        self.end_serial = end_serial
        self.device_names = [device_type + '_' + str(i) for i in range(start_serial, end_serial+1)]
        self.start_len = len(self.device_names)
        self.device_type = device_type

        if device_type == 'TD':

            self.df = pd.DataFrame({'Имя': self.device_names, 
                                    'MAC': [None] * len(self.device_names),
                                    'RSSI': [None] * len(self.device_names),
                                    'Версия прошивки': [None] * len(self.device_names),
                                    'Напряжение батареи': [None] * len(self.device_names),
                                    'Температура': [None] * len(self.device_names),
                                    'Уровень топлива' : [None] * len(self.device_names),
                                    'Период': [None] * len(self.device_names),
                                    'H': [None] * len(self.device_names),
                                    'L': [None] * len(self.device_names),
                                    'Уровень топлива после тарировки': [None] * len(self.device_names),
                                    'Причина отбраковки': [None] * len(self.device_names)})
        elif device_type == 'TH':
        
            self.df = pd.DataFrame({'Имя': self.device_names, 
                                    'MAC': [None] * len(self.device_names),
                                    'RSSI': [None] * len(self.device_names),
                                    'Версия прошивки': [None] * len(self.device_names),
                                    'Напряжение батареи': [None] * len(self.device_names),
                                    'Темпеарутра': [None] * len(self.device_names),
                                    'Влажность': [None] * len(self.device_names),
                                    'Освещенность' : [None] * len(self.device_names)})


    def update_char(self, device_name, char_name, value):
        index = self.df.index[self.df['Имя'] == device_name][0]
        self.df.loc[index, char_name] = value
            

    def get_dataframe(self):
        return self.df
    

    def to_excel(self, xls_path):
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H_%M')
        self.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active

        # определяем цвет заливки
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        white_font = Font(color='FFFFFF') # Белый цвет шрифта

        # вычисляем среднее значение 6 столбца
        avg_temp = np.mean([row[6].value for row in ws.iter_rows(min_row=2)])

        # проходимся по всем строкам, начиная со второй
        for row in ws.iter_rows(min_row=2):
            temp = row[6].value 
            hk = row[9].value 
            lk = row[10].value 
            fl = row[11].value
            if hk and lk and fl is not None:
                if fl > 15:
                    row[12].value = "Уровень топлива более 15 единиц"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif lk < 20000:
                    row[12].value = "L < 20000"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif hk > 43000:
                    row[12].value = "L < 20000"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                elif abs(temp - avg_temp) > 5:
                    row[12].value = "Температура отличается от средней более чем на 5 единиц"
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font


        wb.save(xls_path)
