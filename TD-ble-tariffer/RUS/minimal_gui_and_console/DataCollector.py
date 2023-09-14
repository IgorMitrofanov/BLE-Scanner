import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import datetime
import numpy as np


class DataCollector:
    """
    Описание класса DataCollector.

    Атрибуты:
    - start_serial (int): начальный серийный номер.
    - end_serial (int): конечный серийный номер.
    - device_names (list): список имен устройств.
    - start_len (int): начальная длина списка устройств.
    - device_type (str): тип устройства.
    - TD_length (float): длина ДУТа

    Методы:
    - __init__(self, start_serial, end_serial, device_type): конструктор класса.
    - update_char(self, device_name, char_name, value): метод для обновления значения характеристики.
    - get_dataframe(self): метод для получения данных в виде pandas DataFrame.
    - to_excel(self, xls_path): метод для сохранения данных в файл Excel.
    """
    def __init__(self, start_serial, end_serial, device_type, TD_length):
        """
        Конструктор класса DataCollector.

        Параметры:
        - start_serial (int): начальный серийный номер.
        - end_serial (int): конечный серийный номер.
        - device_type (str): тип устройства.
        """
        self.start_serial = start_serial
        self.end_serial = end_serial
        self.device_names = [device_type + '_' + str(i) for i in range(start_serial, end_serial+1)]
        self.start_len = len(self.device_names)
        self.device_type = device_type
        self.TD_length = TD_length

        if device_type == 'TD':

            self.df = pd.DataFrame({'Имя': self.device_names, 
                                    'MAC': [None] * len(self.device_names),
                                    'RSSI': [None] * len(self.device_names),
                                    'Версия прошивки': [None] * len(self.device_names),
                                    'Напряжение батареи': [None] * len(self.device_names),
                                    'Температура': [None] * len(self.device_names),
                                    'Уровень топлива' : [None] * len(self.device_names),
                                    'Период': [None] * len(self.device_names),
                                    'H': [None] * len(self.device_names),
                                    'L': [None] * len(self.device_names),
                                    'Уровень топлива после тарировки': [None] * len(self.device_names),
                                    'Причина отбраковки': [None] * len(self.device_names)})


    def update_char(self, device_name, char_name, value):
        """
        Метод для обновления значения характеристики.

        Параметры:
        - device_name (str): имя устройства.
        - char_name (str): имя характеристики.
        - value: новое значение характеристики.
        """
        index = self.df.index[self.df['Имя'] == device_name][0]
        self.df.loc[index, char_name] = value
            

    def get_dataframe(self):
        """
        Метод для получения данных в виде pandas DataFrame.

        Возвращает:
        - pandas.DataFrame: данные в виде DataFrame.
        """
        return self.df
    

    def to_excel(self, xls_path, atrribute_error_flag):
        """
        Метод для сохранения данных в файл Excel.

        Параметры:
        - xls_path (str): путь к файлу Excel.
        - atrribute_error_flag (bool): костыль для библиотеки bleak, позволяет сделать пометку, что не все устройства были оттарированы
        """
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H_%M')
        self.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active

        # Определяем цвета заливки
        # Красный - 1ый приоритет ошибки
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        # Оранжевый - 2-ой приоритет ошибки
        orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
        # Желтый - 3-ий приоритет ошибки
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        # Фиолетовый - 4-ый приоритет ошибки
        purple_fill = PatternFill(start_color='FF800080', end_color='FF800080', fill_type='solid')
        # Синий - 5-ый приоритет ошибки
        blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        # Голубой - 6-ой приоритет ошибки
        cyan_fill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')
        # Белый цвет шрифта
        white_font = Font(color='FFFFFF') 
        black_font = Font(color='000000') 

        avg_temp = np.mean([row[6].value for row in ws.iter_rows(min_row=2)])

        # Проходимся по всем строкам, начиная со второй
        for row in ws.iter_rows(min_row=2):
            temp = row[6].value 
            hk = row[9].value 
            lk = row[10].value 
            fl = row[11].value
            b_voltage = row[5].value
            errors_message = ''
            if hk and lk and fl and temp and b_voltage is not None:
                # 6-ой приоритет ошибки
                if abs(temp - avg_temp) > 5:
                    errors_message += "Температура отличается от средней более чем на 5 единиц "
                    for cell in row:
                        cell.fill = cyan_fill 
                        cell.font = white_font
                # 5-ый приоритет ошибки
                elif b_voltage < 3.4:
                    errors_message += "Напряжение батареи меньше порога в 3.4В, проверьте изделие "
                    for cell in row:
                        cell.fill = blue_fill
                        cell.font = white_font
                # 4-ый приоритет ошибки
                elif fl > 15:
                    errors_message.join("Текущий > 15 ")
                    for cell in row:
                        cell.fill = purple_fill
                        cell.font = white_font
                # 3-ий приоритет ошибки
                elif lk < 19000 and self.TD_length == 1:
                    errors_message += "Нижний < 19000 (ДУТ 1 м) "
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = black_font
                elif lk < 28000 and self.TD_length == 1.5:
                    errors_message += "Нижний < 28000 (ДУТ 1.5 м) "
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = black_font
                elif lk < 37000 and self.TD_length == 2:
                    errors_message += "Нижний < 37000 (ДУТ 2 м) "
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = black_font
                elif lk < 46000 and self.TD_length == 3:
                    errors_message += "Нижний < 46000 (ДУТ 3 м) "
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = black_font
                # 2-ой приоритет ошибки
                elif hk > 45000 and self.TD_length == 1:
                    errors_message += "Верхний > 45000 (ДУТ 1 м) "
                    for cell in row:
                        cell.fill = orange_fill
                        cell.font = black_font
                elif hk > 54000 and self.TD_length == 1.5:
                    errors_message += "Верхний > 54000 (ДУТ 1.5 м) "
                    for cell in row:
                        cell.fill = orange_fill
                        cell.font = black_font
                elif hk > 66000 and self.TD_length == 2:
                    errors_message += "Верхний > 66000 (ДУТ 2 м) "
                    for cell in row:
                        cell.fill = orange_fill
                        cell.font = black_font
                elif hk > 77000 and self.TD_length == 3:
                    errors_message += "Верхний > 77000 (ДУТ 3 м) "
                    for cell in row:
                        cell.fill = orange_fill
                        cell.font = black_font
                # 1-ый приоритет ошибки
                elif fl == 7000:
                    errors_message += "Текущий = 7000 "
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = white_font
                row[12].value = errors_message

                
        if atrribute_error_flag == True:
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1).value = "НЕКОТОРЫЕ УСТРОЙСТВА НЕ СМОГИ ЗАВЕРШИТЬ ТАРИРОВКУ, ТРЕБУЕТСЯ ПОВТОРНЫЙ ЗАПУСК ПРОГРАММЫ"
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[last_row]:
                cell.fill = red_fill
                cell.font = font
        wb.save(xls_path)

