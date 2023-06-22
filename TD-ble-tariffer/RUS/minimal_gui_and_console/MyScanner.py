import os

try:
    from bleak import BleakScanner
except:
    os.system('pip install bleak==0.20.2')
    from bleak import BleakScanner

import numpy as np

import asyncio
import random
import datetime
# from queue import Queue

from DataCollector import DataCollector
from adv_decrypt import adv_decrypt
from BleakClientAssistant import BleakClientAssistant

import openpyxl
from openpyxl.styles import PatternFill, Font

try:
    from colorama import init, Fore, Style
except:
    os.system('pip install colorama==0.4.6')

init()

class MyScanner:
    def __init__(self, timeout, start_serial, end_serial, device_type, loop):
        self._scanner = BleakScanner(detection_callback=self.detection_callback, loop=loop)
        self.scanning = asyncio.Event()
        self.timeout = timeout
        self.dc = DataCollector(start_serial, end_serial, device_type)
        self.queue_devices_to_connect = [] 
        self.retry_devices_list = []
        self.loop = loop
        self.device_type = device_type

    async def detection_callback(self, device, advertisement_data):
        try:
            if device.name in self.dc.device_names: 
                
                
                self.queue_devices_to_connect.append(device)
            
                self.dc.device_names.remove(device.name)
            
                print(Fore.GREEN + f'Найдено устройство: {device.name} {self.dc.start_len-len(self.dc.device_names)}/{self.dc.start_len}'+ Style.RESET_ALL)


                self.dc.update_char(device.name, 'MAC', device.address)
                self.dc.update_char(device.name, 'RSSI', advertisement_data.rssi)

                oil_level_raw, battery_voltage, TD_temp_raw, version_raw, cnt_raw = adv_decrypt(advertisement_data.manufacturer_data[3862], device_type=self.device_type)

                self.dc.update_char(device.name, 'Напряжение батареи', battery_voltage)
                self.dc.update_char(device.name, 'Версия прошивки', version_raw)
                self.dc.update_char(device.name, 'Температура', TD_temp_raw)
                self.dc.update_char(device.name, 'Уровень топлива', oil_level_raw)
                self.dc.update_char(device.name, 'Период', cnt_raw)

        except Exception as e:
            #print(Fore.RED + f"Error in callback (scanner): {e}")
            pass

    async def run(self):
        try:
            print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + f'\t\tЗапущенно сканирование в течение {self.timeout} секунд...')
            await self._scanner.start()
            self.scanning.set()
            end_time = self.loop.time() + self.timeout
            while self.scanning.is_set():
                if self.loop.time() > end_time or len(self.dc.device_names) == 0:
                    self.scanning.clear()
                    if len(self.dc.device_names) == 0:
                        print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + '\t\tВсе устройства найдены.')
                        break
                    else:
                        print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.RED + f"\t\t Время сканирования вышло. Не все устройства найдены ({len(self.dc.device_names)*100/self.dc.start_len:.2f}%).")
                        print(Fore.RED + f'Не найденные устройства: {self.dc.device_names}' + Style.RESET_ALL)
                        answer = input(f"Хотите искать снова? (д/любая клавиша для нет): ").lower()
                        if answer == 'д':
                            self.scanning.set()
                            change_timeout = input(f"Хотите изменить время поиска? (д/любая клавиша для нет): ").lower()
                            if change_timeout == 'д':
                                try:
                                    new_timeout = int(input('Время сканирования (только целые числа): '))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + f'\t\tЗапущенно сканирование в течение {self.timeout} секунд...')
                                except:
                                    print(Fore.RED + 'Время сканирования может быть только целочисленное!')
                                    new_timeout = int(input('Время сканирования (только целые числа): '))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + f'\t\tЗапущенно сканирование в течение {self.timeout} секунд...')
                            else:
                                end_time = self.loop.time() + self.timeout
                                print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + f'\t\tЗапущенно сканирование в течение {self.timeout} секунд...')
                        else:
                            print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.RED + f"\t\tВремя сканирования вышло. Не найденные устройства: {self.dc.device_names}")
                else:
                    await asyncio.sleep(0.1)
            await asyncio.sleep(0.1)
            await self._scanner.stop()
            await asyncio.sleep(0.1)
            await self.queue_to_connection()
        except Exception as e:
            #print(Fore.RED + f"Error in run (scanner): {e}")
            pass

    def get_dataframe(self):
        return self.dc.get_dataframe()
    
    def to_excel(self, xls_path):
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H_%M')
        self.dc.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active

        # определяем цвет заливки
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        white_font = Font(color='FFFFFF') # Белый цвет шрифта

        # вычисляем среднее значение 6 столбца
        avg_temp = np.mean([row[6].value for row in ws.iter_rows(min_row=2)])

        # проходимся по всем строкам, начиная со второй
        for row in ws.iter_rows(min_row=2):
            temp = row[6].value 
            hk = row[9].value 
            lk = row[10].value 
            fl = row[11].value
        if fl > 15:
            row[12].value = "Уровень топлива более 15 единиц"
            for cell in row:
                cell.fill = red_fill
                cell.font = white_font
        elif lk < 20000:
            row[12].value = "L < 20000"
            for cell in row:
                cell.fill = red_fill
                cell.font = white_font
        elif hk > 43000:
            row[12].value = "L < 20000"
            for cell in row:
                cell.fill = red_fill
                cell.font = white_font
        elif abs(temp - avg_temp) > 5:
            row[12].value = "Температура отличается от средней более чем на 5 единиц"
            for cell in row:
                cell.fill = red_fill
                cell.font = white_font

        wb.save(xls_path)

    async def connect_device(self, device, loop):
        temp = int(self.dc.df.loc[self.dc.df['Имя'] == device.name, 'Температура'].values[0])
        period = int(self.dc.df.loc[self.dc.df['Имя'] == device.name, 'Период'].values[0])
        fl = int(self.dc.df.loc[self.dc.df['Имя'] == device.name, 'Уровень топлива'].values[0])
        client_assistant = BleakClientAssistant(device, period, temp, fl, loop)
        hk, lk, ul = await client_assistant.run()
        if hk == 0 and lk == 0 and ul == 0:
            self.queue_devices_to_connect.append(device)
            client_assistant = None
            return
        else:
            self.dc.update_char(device.name, 'H', int(hk))
            self.dc.update_char(device.name, 'L', int(lk))
            self.dc.update_char(device.name, 'Уровень топлива после тарировки', int(ul))
            client_assistant = None
            return

    async def queue_to_connection(self):
        while not len(self.queue_devices_to_connect) == 0:
            random_device = random.choice(self.queue_devices_to_connect)
            self.queue_devices_to_connect.remove(random_device)
            print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN + f'\t\tСчетчик успешных тарировок: {self.dc.start_len-len(self.queue_devices_to_connect)-1}/{self.dc.start_len}')
            await self.connect_device(random_device, self.loop)

if __name__ == '__main__':
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    start_serial = int(input('Type start serial (only six numers): '))
    end_serial = int(input('Type start serial (only six numers): '))
    timeout = int(input('Type time in seconds for timeout scanning: '))
    my_scanner = MyScanner(timeout=timeout, start_serial=start_serial, end_serial=end_serial, pattern='TD_', loop=loop)
    loop.run_until_complete(my_scanner.run())
    print(my_scanner.get_dataframe())