import os
from bleak import BleakScanner    
import asyncio
import datetime
from colorama import init, Fore, Style
from DataCollector import DataCollector
from adv_decrypt import adv_decrypt


init()


class MyScanner:
    """
    Initializes the Myscaner object.

    Parameters:
    - timeout: scan timeout in seconds
    - start_serial: initial serial number
    - end_serial: the final serial number
    - device_type: device type
    - loop: asynchronous execution loop
    """ 
    def __init__(self, timeout, start_serial, end_serial, device_type, loop):
        self._scanner = BleakScanner(detection_callback=self.detection_callback, loop=loop)
        self.scanning = asyncio.Event()
        self.timeout = timeout
        self.dc = DataCollector(start_serial, end_serial, device_type)
        self.loop = loop
        self.device_type = device_type


    def detection_callback(self, device, advertisement_data):
        """
        Callback is a function for processing detected devices.

        Parameters:
        - device: detected device
        - advertisement_data: device ad data
        """
        try:
            if device.name in self.dc.device_names: 
                
                
            
                self.dc.device_names.remove(device.name)
            
                print(Fore.GREEN + f'found device with name: {device.name} {self.dc.start_len-len(self.dc.device_names)}/{self.dc.start_len}'+ Style.RESET_ALL)
                self.dc.update_char(device.name, 'MAC', device.address)
                self.dc.update_char(device.name, 'RSSI', advertisement_data.rssi)

                if self.device_type == 'TD':

                    oil_level_raw, battery_voltage, TD_temp_raw, version_raw, cnt_raw = adv_decrypt(advertisement_data.manufacturer_data[3862], device_type=self.device_type)

                    self.dc.update_char(device.name, 'Battery Voltage', battery_voltage)
                    self.dc.update_char(device.name, 'version', version_raw)
                    self.dc.update_char(device.name, 'Temperature', TD_temp_raw)
                    self.dc.update_char(device.name, 'Fuel Level', oil_level_raw)
                    self.dc.update_char(device.name, 'Period', cnt_raw)


                elif self.device_type == 'TH':

                    TH09_temp, TH09_light_raw, TH09_humidity, TH09_battery, TH09_version_raw = adv_decrypt(advertisement_data.manufacturer_data[3862], device_type=self.device_type)

                    self.dc.update_char(device.name, 'Battery Voltage', TH09_battery)
                    self.dc.update_char(device.name, 'version', TH09_version_raw)
                    self.dc.update_char(device.name, 'Temperature', TH09_temp)
                    self.dc.update_char(device.name, 'Humidity', TH09_humidity)
                    self.dc.update_char(device.name, 'Light', TH09_light_raw)

        except Exception as e:
            
            # print(f"Error in callback (scanner): {e}") # Debugging output

            pass
            
            
    async def run(self):
        """
        Starts scanning and collecting data for the specified range of serial numbers and device type.  
        """
        try:
            print(Fore.GREEN  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f'\t\tStarted scanning in {self.device_type}-mode with {self.timeout} seconds timeout...')
            await self._scanner.start()
            self.scanning.set()
            end_time = self.loop.time() + self.timeout
            while self.scanning.is_set():
                if self.loop.time() > end_time or len(self.dc.device_names) == 0:
                    self.scanning.clear()
                    if len(self.dc.device_names) == 0:
                        print(Fore.GREEN  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  '\t\tAll devices have been found.')
                        return
                    else:
                        print(Fore.RED  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f"\t\tTimeout. Can't find all devices ({len(self.dc.device_names)*100/self.dc.start_len:.2f}%)")
                        print(Fore.RED + f'Devices not found: {self.dc.device_names}' + Style.RESET_ALL)
                        answer = input(f"Do you want to scan again? (Y for yes/ANY for no): ").lower()
                        if answer == 'y':
                            self.scanning.set()
                            change_timeout = input(f"Would you like to change timeout? (Y for yes/ANY for no): ").lower()
                            if change_timeout == 'y':
                                try:
                                    new_timeout = int(input('Timeout secdonds (only int): '))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Fore.YELLOW  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f'\t\tStarted scanning with {self.timeout} seconds timeout...')
                                except:
                                    print(Fore.RED + 'timeout can be int only!')
                                    new_timeout = int(input('Timeout secdonds (only int):'))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Fore.YELLOW  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f'\t\tStarted scanning with {self.timeout} seconds timeout...')
                            else:
                                end_time = self.loop.time() + self.timeout
                                print(Fore.YELLOW  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f'\t\tStarted scanning with {self.timeout} seconds timeout...')
                        else:
                            print(Fore.RED  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') +  f"Scan terminated. Devices no found: {self.dc.device_names}")
                else:
                    await asyncio.sleep(0.1)
            await self._scanner.stop()
        except Exception as e:
            
            #print(Fore.RED + f"Error in run (scanner): {e}") # Debugging output
            
            pass


    def get_dataframe(self):
        """
        Returns the Data Frame with the collected data.

        Returns:
        - - Data Frame with collected data
        """
        return self.dc.get_dataframe()
    

    def to_excel(self, xls_path):
        """
        Exports data in Excel format.

        Parameters:
        - xls_path: path to the Excel file
        """
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H-%M')
        self.dc.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            fuel_level = row[7].value 
            period = row[8].value 
            if fuel_level in (6500, 7000):
                row[9].value = "FUEL LEVEL = 6500 OR 7000"
                for cell in row:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF') 
            elif period < 20000:
                row[9].value = "Period < 20000"
                for cell in row:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF') 

        wb.save(xls_path)

