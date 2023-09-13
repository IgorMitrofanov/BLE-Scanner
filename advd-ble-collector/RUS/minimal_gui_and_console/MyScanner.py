import os
from bleak import BleakScanner
from colorama import init, Fore, Style
from DataCollector import DataCollector
from adv_decrypt import adv_decrypt
import datetime
import asyncio
import openpyxl
from openpyxl.styles import PatternFill, Font

init()


class MyScanner:
    """
    Инициализирует объект MyScanner.

    Параметры:
    - timeout: время ожидания сканирования в секундах
    - start_serial: начальный серийный номер
    - end_serial: конечный серийный номер
    - device_type: тип устройства
    - loop: асинхронный цикл исполнения
    """
    def __init__(self, timeout, start_serial, end_serial, device_type, loop):
        self._scanner = BleakScanner(detection_callback=self.detection_callback, loop=loop)
        self.scanning = asyncio.Event()
        self.timeout = timeout
        self.dc = DataCollector(start_serial, end_serial, device_type)
        self.loop = loop
        self.device_type = device_type


    def detection_callback(self, device, advertisement_data):
        """
        Callback-функция для обработки обнаруженных устройств.

        Параметры:
        - device: обнаруженное устройство
        - advertisement_data: данные рекламного объявления устройства
        """
        try:
            if device.name in self.dc.device_names: 
                
                
            
                self.dc.device_names.remove(device.name)
            
                print(Fore.GREEN + f'Найдено устройство: {device.name} {self.dc.start_len-len(self.dc.device_names)}/{self.dc.start_len}'+ Style.RESET_ALL)
                self.dc.update_char(device.name, 'MAC', device.address)
                self.dc.update_char(device.name, 'RSSI', advertisement_data.rssi)

                if self.device_type == 'TD':

                    oil_level_raw, battery_voltage, TD_temp_raw, version_raw, cnt_raw = adv_decrypt(advertisement_data.manufacturer_data[3862], device_type=self.device_type)

                    self.dc.update_char(device.name, 'Напряжение батареи', battery_voltage)
                    self.dc.update_char(device.name, 'Версия прошивки', version_raw)
                    self.dc.update_char(device.name, 'Температура', TD_temp_raw)
                    self.dc.update_char(device.name, 'Уровень топлива', oil_level_raw)
                    self.dc.update_char(device.name, 'Период', cnt_raw)


                elif self.device_type == 'TH':

                    TH09_temp, TH09_light_raw, TH09_humidity, TH09_battery, TH09_version_raw = adv_decrypt(advertisement_data.manufacturer_data[3862], device_type=self.device_type)

                    self.dc.update_char(device.name, 'Напряжение батареи', TH09_battery)
                    self.dc.update_char(device.name, 'Версия прошивки', TH09_version_raw)
                    self.dc.update_char(device.name, 'Температура', TH09_temp)
                    self.dc.update_char(device.name, 'Влажность', TH09_humidity)
                    self.dc.update_char(device.name, 'Освещенность', TH09_light_raw)

        except Exception as e:
            
            # print(f"Error in callback (scanner): {e}") # Отладочный вывод

            pass
            
    async def run(self):
        """
        Запускает сканирование и сбор данных для указанного диапазона серийных номеров и типа устройства.  
        """
        try:
            print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN +  f'\t\tЗапущенно сканирование {self.device_type}-датчиков в течение {self.timeout} секунд...')
            await self._scanner.start()
            self.scanning.set()
            end_time = self.loop.time() + self.timeout
            while self.scanning.is_set():
                if self.loop.time() > end_time or len(self.dc.device_names) == 0:
                    self.scanning.clear()
                    if len(self.dc.device_names) == 0:
                        print(Style.RESET_ALL + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN +  '\t\tВсе устройства найдены.')
                        return
                    else:
                        print(Style.RESET_ALL  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.RED +  f"\t\tВремя сканирования вышло. Не все устройства найдены. ({len(self.dc.device_names)*100/self.dc.start_len:.2f}%)")
                        print(Fore.RED + f'Не найденные устройства: {self.dc.device_names}' + Style.RESET_ALL)
                        answer = input(f"Хотите искать снова? (д/любая клавиша для нет): ").lower()
                        if answer == 'д':
                            self.scanning.set()
                            change_timeout = input(f"Хотите изменить время поиска? (д/любая клавиша для нет): ").lower()
                            if change_timeout == 'д':
                                try:
                                    new_timeout = int(input('Время сканирования (только целые числа): '))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Style.RESET_ALL  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN +  f'\t\tЗапущенно сканирование {self.device_type}-датчиков в течение {self.timeout} секунд...')
                                except:
                                    print(Fore.RED + 'Время сканирования может быть только целочисленное!')
                                    new_timeout = int(input('Время сканирования (только целые числа):'))
                                    self.timeout = new_timeout
                                    end_time = self.loop.time() + self.timeout
                                    print(Style.RESET_ALL  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN +  f'\t\tЗапущенно сканирование {self.device_type}-датчиков в течении {self.timeout} секунд...')
                            else:
                                end_time = self.loop.time() + self.timeout
                                print(Style.RESET_ALL  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.GREEN +  f'\t\tЗапущенно сканирование {self.device_type}-датчиков в течение {self.timeout} секунд...')
                        else:
                            print(Style.RESET_ALL  + datetime.datetime.now().strftime('%Y/%m/%d %H:%M') + Fore.RED +  f"Время сканирования вышло. Не найденные устройства: {self.dc.device_names}")
                else:
                    await asyncio.sleep(0.1)
            await self._scanner.stop()
        except Exception as e:

            #print(Fore.RED + f"Error in run (scanner): {e}") # отладочный вывод

            pass


    def get_dataframe(self):
        """
        Возвращает DataFrame с собранными данными.

        Возвращает:
        - DataFrame с собранными данными
        """
        return self.dc.get_dataframe()
    

    def to_excel(self, xls_path):
        """
        Экспортирует данные в формате Excel.

        Параметры:
        - xls_path: путь к файлу Excel
        """
        date_and_time_write = datetime.datetime.now().strftime('%Y_%m_%d %H-%M')
        self.dc.get_dataframe().to_excel(xls_path, sheet_name=date_and_time_write)

        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active
        if self.device_type == 'TD':
            for row in ws.iter_rows(min_row=2):
                fuel_level = row[7].value 
                period = row[8].value 
                if fuel_level in (6500, 7000):
                    row[9].value = "Уровень топлива = 6500 или 7000"
                    for cell in row:
                        cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFF') 
                elif period < 20000:
                    row[9].value = "Период < 20000"
                    for cell in row:
                        cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFF') 

            wb.save(xls_path)

        elif self.device_type == 'TH':
            for row in ws.iter_rows(min_row=2):
                battarey_voltage = row[5].value 
                if battarey_voltage < 3.4:
                    row[9].value = "Напряжение батареи < 3.4 В, проверьте изделие"
                    for cell in row:
                        cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFF') 

            wb.save(xls_path)